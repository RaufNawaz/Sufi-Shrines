#!/usr/bin/env python3
"""
shrines_enrich.py — incremental enrichment engine for Shrines_with_Descriptions.xlsx

This is the SAFE, DETERMINISTIC half of the enrichment workflow. The research +
writing half is done by Claude following ENRICHMENT_RUNBOOK.md. This script:

  --status            Show current gaps (missing descriptions / images) and the
                      next candidate sites from _enrichment_queue.md.
  --write BATCH.md    Parse a staging batch file (same delimited format the
                      runbook produces), BACK UP the workbook, then:
                        * fill Description cells for <<<ENTRY>>> blocks
                          (only if empty; verifies the row's Name still matches),
                        * append new rows for <<<SHRINE add=YES>>> blocks
                          (skips duplicates; copies cell styles; adds separator),
                      then append a record to _ENRICHMENT_LOG.md.

Never overwrites an existing description. Never edits without a fresh backup.

Usage:
  python3 tools/shrines_enrich.py --status [--limit 15]
  python3 tools/shrines_enrich.py --write _enrichment_batch.md
"""
import argparse, re, shutil, sys, io, datetime
from pathlib import Path
import openpyxl
from copy import copy

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

PROJECT_DIR = Path(__file__).resolve().parent.parent
XLSX   = PROJECT_DIR / "Shrines_with_Descriptions.xlsx"
QUEUE  = PROJECT_DIR / "_enrichment_queue.md"
LOG    = PROJECT_DIR / "_ENRICHMENT_LOG.md"
BACKUP_DIR = PROJECT_DIR / "archive" / "xlsx_backups"
SEP = "\n\n" + "=" * 85
COLS = ["Name","Location","Category","Latitude","Longitude",
        "Founded/Opened","Sufi Saint","Image 1","Image 2","Events","Description"]

def _has(v): return v is not None and str(v).strip() != ""

def _norm(name):
    """Normalise a site name for duplicate detection."""
    s = (name or "").lower()
    s = re.sub(r"\(.*?\)", " ", s)                     # drop parentheticals
    s = re.sub(r"^(shrine|tomb|mausoleum|roza|dargah|mazar)\s+of\s+", " ", s)
    s = re.sub(r"\b(shrine|tomb|mausoleum|dargah|darbar|roza|mazar|gurudwara|"
               r"gurdwara|mandir|temple|hazrat|the|of)\b", " ", s)
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()

def load(data_only=False):
    if not XLSX.exists():
        sys.exit(f"ERROR: {XLSX} not found.")
    wb = openpyxl.load_workbook(XLSX, data_only=data_only)
    ws = wb["Shrines"] if "Shrines" in wb.sheetnames else wb.active
    header = [c.value for c in ws[1]]
    col = {h: i + 1 for i, h in enumerate(header)}
    return wb, ws, col

def backup():
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.datetime.now().strftime("%Y-%m-%d_%H%M%S")
    dest = BACKUP_DIR / f"Shrines_with_Descriptions.{ts}.xlsx"
    shutil.copy2(XLSX, dest)
    return dest

# ---------------- status ----------------
def cmd_status(limit):
    wb, ws, col = load(data_only=True)
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    ci = {k: col[k] - 1 for k in col}
    miss_desc = [(i + 2, r[ci["Name"]], r[ci["Category"]])
                 for i, r in enumerate(rows) if not _has(r[ci["Description"]])]
    miss_img  = [(i + 2, r[ci["Name"]], r[ci["Category"]])
                 for i, r in enumerate(rows) if not _has(r[ci["Image 1"]])]
    print(f"Workbook: {XLSX.name}")
    print(f"Data rows: {len(rows)}  |  missing Description: {len(miss_desc)}  |  missing Image 1: {len(miss_img)}")

    def show(title, items):
        print(f"\n{title} (showing up to {limit}):")
        for r, n, c in items[:limit]:
            print(f"  row {r:>3} | {c or '?':<14} | {n}")
        if len(items) > limit:
            print(f"  ... and {len(items) - limit} more")

    show("ROWS MISSING A DESCRIPTION", miss_desc)
    show("ROWS MISSING Image 1", miss_img)

    # queue preview
    existing = {_norm(r[ci["Name"]]) for r in rows if _has(r[ci["Name"]])}
    if QUEUE.exists():
        pending = []
        for line in QUEUE.read_text(encoding="utf-8").splitlines():
            m = re.match(r"\s*-\s*\[ \]\s*(.+)", line)
            if m:
                name = m.group(1).split("—")[0].split("|")[0].strip()
                flag = "  (ALREADY IN SHEET?)" if _norm(name) in existing else ""
                pending.append(m.group(1).strip() + flag)
        print(f"\nNEXT CANDIDATE SITES TO ADD (from {QUEUE.name}, {len(pending)} pending, showing {limit}):")
        for p in pending[:limit]:
            print(f"  - {p}")
        if not pending:
            print("  (queue empty — the runbook says to research & append new candidates)")
    else:
        print(f"\n(no {QUEUE.name} yet — create one or let the runbook seed it)")
    print("\nNext: pick a batch, write it to a staging .md, then run --write on it.")

# ---------------- parsing ----------------
def parse_entries(text):
    return [(int(r), name, body.strip())
            for r, name, body in re.findall(
                r"<<<ENTRY row=(\d+) name=\"([^\"]*)\">>>\n(.*?)\n<<<END>>>", text, re.DOTALL)]

def parse_shrines(text):
    out = []
    for block in re.findall(r"<<<SHRINE add=YES>>>\n(.*?)\n<<<END>>>", text, re.DOTALL):
        d = {}
        for line in block.splitlines():
            m = re.match(r"^(NAME|LOCATION|CATEGORY|LATITUDE|LONGITUDE|FOUNDED|SAINT|"
                         r"IMAGE1|IMAGE2|EVENTS|DESCRIPTION|NOTE):\s?(.*)$", line)
            if m: d[m.group(1)] = m.group(2).strip()
        if d.get("NAME"): out.append(d)
    return out

def _clean(v): return "" if v is None or v.strip().upper() == "NONE" else v.strip()

# ---------------- write ----------------
def cmd_write(batch_path):
    p = Path(batch_path)
    if not p.is_absolute(): p = PROJECT_DIR / batch_path
    if not p.exists(): sys.exit(f"ERROR: batch file {p} not found.")
    text = p.read_text(encoding="utf-8")
    entries, shrines = parse_entries(text), parse_shrines(text)
    if not entries and not shrines:
        sys.exit("ERROR: no <<<ENTRY>>> or <<<SHRINE add=YES>>> blocks found in batch.")

    bpath = backup()
    wb, ws, col = load()
    existing_names = {_norm(ws.cell(row=r, column=col["Name"]).value)
                      for r in range(2, ws.max_row + 1)
                      if _has(ws.cell(row=r, column=col["Name"]).value)}

    log = [f"\n## Enrichment run {datetime.datetime.now():%Y-%m-%d %H:%M:%S}",
           f"- Backup: `{bpath.relative_to(PROJECT_DIR)}`"]

    # 1) descriptions
    desc_done, desc_skip = [], []
    for rownum, name, body in entries:
        cell = ws.cell(row=rownum, column=col["Description"])
        cur_name = ws.cell(row=rownum, column=col["Name"]).value
        if name and _norm(cur_name) != _norm(name):
            desc_skip.append(f"row {rownum}: NAME MISMATCH (sheet='{cur_name}' vs batch='{name}') — skipped")
            continue
        if _has(cell.value):
            desc_skip.append(f"row {rownum} ({cur_name}): already has a description — skipped")
            continue
        cell.value = body + SEP
        desc_done.append(f"row {rownum}: {cur_name} ({len(cell.value)} chars)")

    # 2) new rows
    src = {c: ws.cell(row=2, column=col[c]) for c in COLS}
    added, dup = [], []
    r = ws.max_row + 1
    for d in shrines:
        if _norm(d["NAME"]) in existing_names:
            dup.append(f"{d['NAME']} — duplicate, skipped"); continue
        vals = {
            "Name": d.get("NAME",""), "Location": d.get("LOCATION",""),
            "Category": d.get("CATEGORY",""), "Latitude": _clean(d.get("LATITUDE","")),
            "Longitude": _clean(d.get("LONGITUDE","")), "Founded/Opened": d.get("FOUNDED",""),
            "Sufi Saint": d.get("SAINT",""), "Image 1": _clean(d.get("IMAGE1","")),
            "Image 2": _clean(d.get("IMAGE2","")), "Events": d.get("EVENTS",""),
            "Description": (d["DESCRIPTION"].strip() + SEP) if d.get("DESCRIPTION","").strip() else "",
        }
        for c in COLS:
            cell = ws.cell(row=r, column=col[c]); cell.value = vals[c]
            s = src[c]
            cell.font = copy(s.font); cell.alignment = copy(s.alignment)
            cell.number_format = s.number_format; cell.border = copy(s.border); cell.fill = copy(s.fill)
        existing_names.add(_norm(d["NAME"]))
        added.append(f"row {r}: {vals['Name']} [{vals['Category']}] img1={'Y' if vals['Image 1'] else 'N'}")
        r += 1

    wb.save(XLSX)

    # report + log
    print(f"Backup: {bpath}")
    print(f"Descriptions filled: {len(desc_done)} | skipped: {len(desc_skip)}")
    for x in desc_done: print("  +", x)
    for x in desc_skip: print("  ~", x)
    print(f"New rows added: {len(added)} | duplicates skipped: {len(dup)}")
    for x in added: print("  +", x)
    for x in dup: print("  ~", x)
    print(f"New max_row: {ws.max_row}")

    log.append(f"- Descriptions filled: {len(desc_done)} " +
               (f"({'; '.join(desc_done)})" if desc_done else ""))
    if desc_skip: log.append(f"- Description skips: {'; '.join(desc_skip)}")
    log.append(f"- New rows added: {len(added)} " + (f"({'; '.join(added)})" if added else ""))
    if dup: log.append(f"- Duplicates skipped: {'; '.join(dup)}")
    log.append(f"- Workbook now has {ws.max_row - 1} data rows.")
    with open(LOG, "a", encoding="utf-8") as f: f.write("\n".join(log) + "\n")
    print(f"Logged to {LOG.name}")

def main():
    ap = argparse.ArgumentParser(description="Incremental enrichment engine for the shrines spreadsheet.")
    ap.add_argument("--status", action="store_true", help="show gaps + next queue candidates")
    ap.add_argument("--write", metavar="BATCH.md", help="write a staging batch into the workbook")
    ap.add_argument("--limit", type=int, default=15, help="rows to show in --status (default 15)")
    a = ap.parse_args()
    if a.write: cmd_write(a.write)
    elif a.status: cmd_status(a.limit)
    else: ap.print_help()

if __name__ == "__main__":
    main()
